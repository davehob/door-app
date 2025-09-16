from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.table import Table, TableStyleInfo

def build_order_excel(order, items):
    wb = Workbook()

    # Summary
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "Door Order"; ws["A1"].font = Font(size=16, bold=True)
    ws["A3"] = "Order ID"; ws["B3"] = order.id
    ws["A4"] = "Job ID"; ws["B4"] = order.job_id
    ws["A5"] = "Dealer Code"; ws["B5"] = order.dealer_code or ""
    ws["A6"] = "Job Name"; ws["B6"] = order.job_name or ""
    ws["A7"] = "Finish"; ws["B7"] = order.finish or ""
    ws["A9"]  = "Door SFP"; ws["B9"]  = order.style_door_sfp or ""
    ws["A10"] = "Door Flat"; ws["B10"] = order.style_door_flat or ""
    ws["A11"] = "Drawer SFP"; ws["B11"] = order.style_drawer_sfp or ""
    ws["A12"] = "Drawer Flat"; ws["B12"] = order.style_drawer_flat or ""
    ws["A13"] = "Panel Code"; ws["B13"] = order.style_panel_code or ""
    ws["A15"] = "Hinge Top Offset (in)"; ws["B15"] = order.hinge_top_offset_in or ""
    ws["A16"] = "Hinge Bottom Offset (in)"; ws["B16"] = order.hinge_bottom_offset_in or ""
    ws["A17"] = "Hinge Size (in)"; ws["B17"] = order.hinge_size_in or ""

    # Items
    ws2 = wb.create_sheet("Items")
    headers = ["Line","Type","Style","Qty","Width (in)","Height (in)","Hinge","Notes/Hinge","Source Page"]
    ws2.append(headers)
    for i, it in enumerate(items, start=1):
        ws2.append([
            i, it.type, it.style, it.qty,
            it.width_in, it.height_in,
            it.hinge or "None",
            it.note, it.source_page
        ])

    ref = f"A1:I{len(items)+1}"
    tbl = Table(displayName="tbl_items", ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws2.add_table(tbl)
    for col in "ABCDEFGHI":
        ws2.column_dimensions[col].width = 15
    ws2.freeze_panes = "A2"
    return wb